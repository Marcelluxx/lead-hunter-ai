"""
Lead Hunter V3 — Excel Exporter (openpyxl)
Export duale:
  - Modalità "no_website": colonne base (senza ideal_product/sales_hook)
  - Modalità "with_website": colonne estese con audit sito web
"""

import os
from typing import Union, List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from .filters import clean_and_translate_categories, extract_address_details
except (ImportError, ValueError):
    from filters import clean_and_translate_categories, extract_address_details



class DataExporter:
    """Esportatore professionale con formattazione avanzata openpyxl."""

    # Stili header premium
    HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
    HEADER_BORDER = Border(
        bottom=Side(style="medium", color="64748B"),
        right=Side(style="thin", color="334155"),
    )

    # Stili celle dati
    DATA_FONT = Font(name="Calibri", size=10)
    DATA_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
    ALT_ROW_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    # Colori score condizionale
    SCORE_GREEN = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    SCORE_YELLOW = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
    SCORE_RED = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

    @staticmethod
    def export_to_excel(
        leads: Union[List[Dict], Dict[str, Dict]],
        mode: str = "no_website",
        filename: str = "leads_v3_premium.xlsx"
    ) -> None:
        """
        Esporta lead in Excel con formattazione professionale.
        mode: "no_website" | "with_website"
        """
        if not leads:
            print("[Exporter] Nessun lead da esportare.")
            return

        # Assicurati che la cartella di destinazione esista
        dir_name = os.path.dirname(filename)
        if dir_name:
            os.makedirs(dir_name, exist_ok=True)

        leads_list = list(leads.values()) if isinstance(leads, dict) else leads

        # Definisci colonne in base alla modalità
        if mode == "with_website":
            columns = DataExporter._get_website_columns()
            rows = DataExporter._format_website_rows(leads_list)
        else:
            columns = DataExporter._get_no_website_columns()
            rows = DataExporter._format_no_website_rows(leads_list)

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Leads"

            # --- HEADER ROW ---
            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = DataExporter.HEADER_FONT
                cell.fill = DataExporter.HEADER_FILL
                cell.alignment = DataExporter.HEADER_ALIGNMENT
                cell.border = DataExporter.HEADER_BORDER

            # Altezza header
            ws.row_dimensions[1].height = 30

            # --- DATA ROWS ---
            for row_idx, row_data in enumerate(rows, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = DataExporter.DATA_FONT
                    cell.alignment = DataExporter.DATA_ALIGNMENT

                    # Righe alternate
                    if row_idx % 2 == 0:
                        cell.fill = DataExporter.ALT_ROW_FILL

                # Colorazione condizionale per Website Score
                if mode == "with_website":
                    score_col = columns.index("Website Score") + 1 if "Website Score" in columns else None
                    if score_col:
                        score_cell = ws.cell(row=row_idx, column=score_col)
                        try:
                            score_val = int(score_cell.value) if score_cell.value else 0
                            if score_val <= 3:
                                score_cell.fill = DataExporter.SCORE_RED
                            elif score_val <= 6:
                                score_cell.fill = DataExporter.SCORE_YELLOW
                            else:
                                score_cell.fill = DataExporter.SCORE_GREEN
                        except (ValueError, TypeError):
                            pass

            # --- AUTO-FIT COLONNE ---
            col_widths = DataExporter._calculate_column_widths(columns, rows)
            for col_idx, width in enumerate(col_widths, 1):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = width

            # Freeze header row
            ws.freeze_panes = "A2"

            # Auto-filter
            ws.auto_filter.ref = ws.dimensions

            wb.save(filename)
            print(f"\n✅ [OK] Esportazione premium completata: {len(leads_list)} lead in '{filename}'")

        except PermissionError:
            print(f"❌ [Errore] Il file '{filename}' è aperto in un altro programma. Chiudilo e riprova.")
        except Exception as e:
            print(f"❌ [Errore] Eccezione durante l'esportazione: {e}")

    @staticmethod
    def _get_no_website_columns() -> list:
        return [
            "Business Name", "Category", "Address", "Paese", "Phone",
            "Rating", "Reviews", "Top Competitor",
            "Business Summary", "Key Weakness"
        ]

    @staticmethod
    def _get_website_columns() -> list:
        return [
            "Business Name", "Category", "Address", "Paese", "Phone",
            "Rating", "Reviews", "Website",
            "Extracted Email", "Website Score",
            "Diagnosis", "Site Brief", "Cold Message"
        ]

    @staticmethod
    def _format_no_website_rows(leads: List[Dict]) -> List[list]:
        rows = []
        for lead in leads:
            category = clean_and_translate_categories(lead.get("types", []), lead.get("search_keyword", ""))
            via_e_civico, paese = extract_address_details(lead)

            rows.append([
                lead.get("displayName", {}).get("text", "N/A"),
                category,
                via_e_civico,
                paese,
                lead.get("nationalPhoneNumber", "N/A"),
                lead.get("rating", "N/A"),
                lead.get("userRatingCount", "N/A"),
                lead.get("competitor", "N/A"),
                lead.get("business_summary", "N/A"),
                lead.get("key_weakness", "N/A"),
            ])
        return rows

    @staticmethod
    def _format_website_rows(leads: List[Dict]) -> List[list]:
        rows = []
        for lead in leads:
            category = clean_and_translate_categories(lead.get("types", []), lead.get("search_keyword", ""))
            via_e_civico, paese = extract_address_details(lead)

            # Email: preferisci quella estratta dal crawler, fallback su Google Places
            extracted_email = lead.get("extracted_email", "")
            if isinstance(extracted_email, list):
                extracted_email = ", ".join(extracted_email) if extracted_email else ""

            rows.append([
                lead.get("displayName", {}).get("text", "N/A"),
                category,
                via_e_civico,
                paese,
                lead.get("nationalPhoneNumber", "N/A"),
                lead.get("rating", "N/A"),
                lead.get("userRatingCount", "N/A"),
                lead.get("websiteUri", "N/A"),
                extracted_email or "N/A",
                lead.get("website_score", "N/A"),
                lead.get("diagnosis", "N/A"),
                lead.get("site_brief", "N/A"),
                lead.get("cold_message", "N/A"),
            ])
        return rows

    @staticmethod
    def _calculate_column_widths(columns: list, rows: list) -> list:
        """Calcola larghezze ottimali per ogni colonna."""
        widths = []
        for col_idx, col_name in enumerate(columns):
            max_len = len(col_name)
            for row in rows:
                if col_idx < len(row):
                    cell_len = len(str(row[col_idx] or ""))
                    max_len = max(max_len, cell_len)

            # Limiti per colonne lunghe (diagnosis, cold_message)
            if col_name in ("Diagnosis", "Cold Message", "Site Brief"):
                widths.append(min(max_len + 2, 50))
            elif col_name in ("Business Summary", "Key Weakness", "Address"):
                widths.append(min(max_len + 2, 40))
            else:
                widths.append(min(max_len + 2, 30))
        return widths
